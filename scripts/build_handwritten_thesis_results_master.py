#!/usr/bin/env python3
"""Build the thesis master Excel workbook for handwritten experiments."""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


THESIS_PREDICTIONS_ROOT = Path(
    "/Users/dennisberger/Library/Mobile Documents/com~apple~CloudDocs/Dokumente/Uni/Master_Thesis/predictions"
)
DEFAULT_OUTPUT = THESIS_PREDICTIONS_ROOT / "handwritten_thesis_results_master.xlsx"
THESIS_BUNDLE = THESIS_PREDICTIONS_ROOT / "handwritten_thesis_bundle_2026-04-10"
THESIS_RUN_ROOT = Path(
    "/Users/dennisberger/Documents/Privat/llm-line-alignment/cluster_runs/thesis_handwritten_completion_2026-04-09"
)
CLUSTER_RESULTS_MAR31 = THESIS_PREDICTIONS_ROOT / "cluster_results_2026-03-31_15-10-05"
OLDER_BUCKETS = {
    "legacy_jan_all_methods": THESIS_PREDICTIONS_ROOT / "2026_01_14_evalualation_all_methods",
    "legacy_children_jan21": THESIS_PREDICTIONS_ROOT / "2026_01_21_evaluation_children_hw",
    "legacy_gpt52_feb11": THESIS_PREDICTIONS_ROOT / "2026_02_11_evaluation_gpt5-2",
    "legacy_mistral_feb16": THESIS_PREDICTIONS_ROOT / "2026_02_16_evaluation_mistral-large",
    "legacy_gemini_mar02": THESIS_PREDICTIONS_ROOT / "2026_03_02_evaluation_gemini-3-pro",
    "legacy_nntp_mar16": THESIS_PREDICTIONS_ROOT / "2026_03_16_nntp_first_run",
    "legacy_nntp_mar20": THESIS_PREDICTIONS_ROOT / "2026_03_20_nntp_cluster_results",
}
BULLINGER_SUBSETS_MANIFEST = Path("datasets/bullinger_handwritten/subsets/manifest.json")
CORE_DATASETS = {
    "bullinger_handwritten",
    "children_handwritten",
    "washington_handwritten",
    "iam_handwritten_rwth_test_representative_20",
}
EXPECTED_COUNTS = {
    "bullinger_handwritten": 59,
    "children_handwritten": 63,
    "washington_handwritten": 20,
    "iam_handwritten_rwth_test_representative_20": 20,
}
METRIC_COLUMNS = ["wer", "cer", "line_acc", "rev_line_acc", "exact_line_f1"]
MASTER_COLUMNS = [
    "scope",
    "dataset",
    "dataset_display",
    "dataset_group",
    "subset_group",
    "n_samples",
    "method",
    "method_variant",
    "model",
    "model_family",
    "provider",
    "shots",
    "baseline_type",
    "ocr_status",
    "source_bucket",
    "source_csv",
    "run_date_bucket",
    "thesis_status",
    "wer",
    "cer",
    "line_acc",
    "rev_line_acc",
    "exact_line_f1",
    "notes",
]
SUBSET_COLUMNS = [
    "scope",
    "dataset",
    "dataset_display",
    "subset_group",
    "subset_label",
    "n_samples",
    "method",
    "method_variant",
    "model",
    "baseline_type",
    "source_bucket",
    "source_csv",
    "thesis_status",
    "wer",
    "cer",
    "line_acc",
    "rev_line_acc",
    "exact_line_f1",
    "notes",
]
AUDIT_COLUMNS = [
    "source_bucket",
    "source_csv",
    "dataset",
    "method",
    "method_variant",
    "model",
    "n_samples",
    "decision",
    "reason",
    "scope",
    "notes",
]
SUMMARY_SHEET_NAME = "technical_summary"
DATASET_ORDER = [
    "bullinger_handwritten",
    "children_handwritten",
    "iam_handwritten_rwth_test_representative_20",
    "washington_handwritten",
]
METHOD_ORDER = ["m1", "m2", "m3", "m4", "m5"]
DATASET_SUMMARY_COLUMNS = [
    "dataset",
    "n_samples",
    "nntp_line_acc",
    "nntp_provenance",
    "m1_max_line_acc",
    "m2_max_line_acc",
    "m3_max_line_acc",
    "m4_max_line_acc",
    "m5_max_line_acc",
    "best_method",
    "best_model",
    "best_line_acc",
    "delta_best_vs_nntp",
]
METHOD_SUMMARY_COLUMNS = [
    "method",
    "best_model",
    "mean_line_acc",
    "datasets_covered",
    "best_dataset",
    "best_dataset_line_acc",
]
BULLINGER_SUMMARY_COLUMNS = [
    "subset",
    "local_n_samples",
    "nntp_paper_line_acc",
    "m1_max_line_acc",
    "m2_max_line_acc",
    "m3_max_line_acc",
    "m4_max_line_acc",
    "m5_max_line_acc",
    "best_method",
    "best_model",
    "best_line_acc",
    "delta_best_vs_nntp",
]

SUMMARY_ROW_IDS = {
    "",
    "macro_avg",
    "micro_avg",
    "fold_a",
    "fold_b",
    "fold_c",
    "train_a_test_b",
    "train_b_test_a",
}
PAPER_URL = "https://arxiv.org/pdf/2508.07904"
PAPER_BULLINGER = {
    "Subset1": {"cer": 0.076, "wer": 0.289, "line_acc": 0.861},
    "Subset2": {"cer": 0.186, "wer": 0.540, "line_acc": 0.773},
    "line_counts": {"Subset1": 902, "Subset2": 1486},
}


@dataclass
class Candidate:
    source_bucket: str
    source_csv: str
    source_path: Path | None
    dataset: str
    dataset_display: str
    dataset_group: str
    method: str
    method_variant: str
    model: str
    model_family: str
    provider: str
    shots: int | None
    baseline_type: str
    ocr_status: str
    run_date_bucket: str
    n_samples: int | None
    metrics: dict[str, float | None]
    notes: str
    scope: str | None = None
    thesis_status: str | None = None

    @property
    def row_key(self) -> tuple[str, str, str, str]:
        return (self.dataset, self.method, self.model, self.method_variant)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--out",
        default=str(DEFAULT_OUTPUT),
        help="Output XLSX path.",
    )
    return ap.parse_args()


def normalize_dataset(raw: str) -> str:
    mapping = {
        "iam_repr20": "iam_handwritten_rwth_test_representative_20",
        "iam_handwritten_rwth_test_representative_20": "iam_handwritten_rwth_test_representative_20",
        "easy_hist": "easy_historical",
    }
    return mapping.get(raw, raw)


def dataset_display(dataset: str) -> str:
    mapping = {
        "bullinger_handwritten": "Bullinger Handwritten",
        "children_handwritten": "Children Handwritten",
        "washington_handwritten": "Washington Handwritten",
        "iam_handwritten_rwth_test_representative_20": "IAM Handwritten RWTH Test Representative 20",
        "iam_handwritten": "IAM Handwritten",
        "bullinger_print": "Bullinger Print",
        "iam_print": "IAM Print",
        "easy_historical": "Easy Historical",
    }
    return mapping.get(dataset, dataset.replace("_", " ").title())


def dataset_group(dataset: str) -> str:
    if dataset in CORE_DATASETS:
        return "handwritten_core"
    if dataset.endswith("_print"):
        return "print"
    if dataset in {"easy_historical"}:
        return "historical"
    if "handwritten" in dataset:
        return "handwritten_appendix"
    return "other"


def model_info(model: str) -> tuple[str, str]:
    m = model.lower()
    if m.startswith("gpt-"):
        return "openai", "gpt"
    if m.startswith("gemini-"):
        return "google", "gemini"
    if m.startswith("mistral-"):
        return "mistral", "mistral"
    if "qwen" in m:
        return "huggingface", "qwen"
    if "nntp" in m:
        return "paper" if "paper" in m else "local", "nntp"
    if m == "legacy_unspecified":
        return "unknown", "legacy_unspecified"
    return "unknown", model


def float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    return float(value)


def read_eval_df(path: Path) -> pd.DataFrame:
    return pd.read_csv(path)


def macro_metrics(df: pd.DataFrame) -> dict[str, float | None]:
    if "id" in df.columns and (df["id"].astype(str) == "macro_avg").any():
        row = df.loc[df["id"].astype(str) == "macro_avg"].iloc[0]
    else:
        row = df.iloc[-1]
    return {metric: float_or_none(row.get(metric)) for metric in METRIC_COLUMNS}


def sample_rows(df: pd.DataFrame, dataset: str) -> pd.DataFrame:
    if "id" not in df.columns:
        return df.iloc[0:0]
    ids = df["id"].astype(str).fillna("")
    mask = ~ids.isin(SUMMARY_ROW_IDS) & ids.ne("nan")
    out = df.loc[mask].copy()
    out = out[out["id"].notna()]
    if dataset == "bullinger_handwritten":
        subset_ids = set(load_bullinger_subset_map())
        out = out[out["id"].astype(str).isin(subset_ids)]
    return out


def detect_n_samples(df: pd.DataFrame, dataset: str, method: str, method_variant: str) -> int | None:
    if dataset in EXPECTED_COUNTS and method == "nntp" and method_variant in {"cv_macro", "legacy_single_run"}:
        return EXPECTED_COUNTS[dataset]
    rows = sample_rows(df, dataset)
    if not rows.empty:
        return int(len(rows))
    return None


def infer_ocr_status(dataset: str, method: str, source_bucket: str, method_variant: str) -> str:
    if method == "m1":
        return "no_ocr"
    if method == "nntp":
        return "external_paper" if source_bucket == "external_paper_2508.07904" else "pylaia_ctc"
    if dataset == "children_handwritten" and method in {"m2", "m3", "m4"} and source_bucket in {
        "thesis_bundle_2026-04-10",
        "thesis_run_root_2026-04-09",
    }:
        return "children_fixed_apr03"
    if dataset == "children_handwritten" and method == "m5" and source_bucket in {
        "thesis_bundle_2026-04-10",
        "thesis_run_root_2026-04-09",
    }:
        return "children_fixed_apr03"
    if method == "m5" and method_variant == "context100":
        return "ocr_text_context100"
    return "standard_ocr"


def candidate_from_df(
    *,
    source_bucket: str,
    source_path: Path,
    dataset: str,
    method: str,
    method_variant: str,
    model: str,
    shots: int | None,
    baseline_type: str,
    run_date_bucket: str,
    notes: str = "",
) -> Candidate:
    dataset = normalize_dataset(dataset)
    df = read_eval_df(source_path)
    provider, model_family = model_info(model)
    return Candidate(
        source_bucket=source_bucket,
        source_csv=str(source_path),
        source_path=source_path,
        dataset=dataset,
        dataset_display=dataset_display(dataset),
        dataset_group=dataset_group(dataset),
        method=method,
        method_variant=method_variant,
        model=model,
        model_family=model_family,
        provider=provider,
        shots=shots,
        baseline_type=baseline_type,
        ocr_status=infer_ocr_status(dataset, method, source_bucket, method_variant),
        run_date_bucket=run_date_bucket,
        n_samples=detect_n_samples(df, dataset, method, method_variant),
        metrics=macro_metrics(df),
        notes=notes,
    )


def load_bullinger_subset_manifest() -> dict[str, Any]:
    return json.loads(BULLINGER_SUBSETS_MANIFEST.read_text(encoding="utf-8"))


def load_bullinger_subset_map() -> dict[str, str]:
    manifest = load_bullinger_subset_manifest()
    subset_map: dict[str, str] = {}
    for subset_name, ids in manifest["subsets"].items():
        for sample_id in ids:
            subset_map[sample_id] = subset_name
    return subset_map


def collect_bundle_candidates() -> list[Candidate]:
    candidates: list[Candidate] = []
    for csv_path in sorted(THESIS_BUNDLE.rglob("evaluation.csv")):
        rel = csv_path.relative_to(THESIS_BUNDLE)
        parts = rel.parts
        if len(parts) != 4:
            continue
        dataset, method_dir, model, _ = parts
        if method_dir == "m5_context100":
            method = "m5"
            method_variant = "context100"
            shots = 0
            baseline_type = "llm"
        elif method_dir in {"m2", "m3", "m4"}:
            method = method_dir
            method_variant = "0shot"
            shots = 0
            baseline_type = "llm"
        else:
            continue
        candidates.append(
            candidate_from_df(
                source_bucket="thesis_bundle_2026-04-10",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=method_variant,
                model=model,
                shots=shots,
                baseline_type=baseline_type,
                run_date_bucket="2026-04-10_bundle",
            )
        )
    return candidates


def collect_runroot_candidates() -> list[Candidate]:
    candidates: list[Candidate] = []
    for csv_path in sorted(THESIS_RUN_ROOT.rglob("evaluation.csv")):
        rel = csv_path.relative_to(THESIS_RUN_ROOT)
        parts = rel.parts
        if len(parts) != 4:
            continue
        dataset, method_dir, model, _ = parts
        if method_dir == "m5_context100":
            method = "m5"
            method_variant = "context100"
            shots = 0
        elif method_dir in {"m2", "m3", "m4"}:
            method = method_dir
            method_variant = "0shot"
            shots = 0
        else:
            continue
        candidates.append(
            candidate_from_df(
                source_bucket="thesis_run_root_2026-04-09",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=method_variant,
                model=model,
                shots=shots,
                baseline_type="llm",
                run_date_bucket="2026-04-09_runroot",
            )
        )
    nntp_path = THESIS_RUN_ROOT / "children_handwritten" / "nntp" / "children_handwritten_eval_nntp_cv.csv"
    if nntp_path.exists():
        candidates.append(
            candidate_from_df(
                source_bucket="thesis_run_root_2026-04-09",
                source_path=nntp_path,
                dataset="children_handwritten",
                method="nntp",
                method_variant="cv_macro",
                model="NNTP",
                shots=None,
                baseline_type="nntp",
                run_date_bucket="2026-04-09_runroot",
            )
        )
    return candidates


def parse_cluster_eval_name(name: str) -> tuple[str, str, str, str, int | None] | None:
    nntp_special = re.fullmatch(r"(.+)_eval_nntp_(cv_macro|train_a_test_b|train_b_test_a|iam_zero_adapt)\.csv", name)
    if nntp_special:
        dataset = normalize_dataset(nntp_special.group(1))
        tag = nntp_special.group(2)
        variant = "cv_macro" if tag == "cv_macro" else tag
        return dataset, "nntp", variant, "NNTP", None

    legacy_nntp = re.fullmatch(r"(.+)_eval_nntp\.csv", name)
    if legacy_nntp:
        dataset = normalize_dataset(legacy_nntp.group(1))
        return dataset, "nntp", "legacy_single_run", "NNTP", None

    pattern = re.fullmatch(
        r"(.+)_eval_(m[1-4]|m5_separate)_(0shot|1shot)(?:_(prompt_[^.]+))?(?:_(.+))?\.csv",
        name,
    )
    if pattern:
        dataset = normalize_dataset(pattern.group(1))
        raw_method = pattern.group(2)
        shot = pattern.group(3)
        prompt_variant = pattern.group(4)
        model = pattern.group(5) or "legacy_unspecified"
        method = "m5" if raw_method == "m5_separate" else raw_method
        variant = "separate_0shot" if raw_method == "m5_separate" else shot
        if prompt_variant:
            variant = f"{shot}_{prompt_variant}"
        return dataset, method, variant, model, int(shot[0])

    return None


def collect_cluster_results_candidates() -> list[Candidate]:
    candidates: list[Candidate] = []
    for csv_path in sorted(CLUSTER_RESULTS_MAR31.glob("*.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            if csv_path.name.startswith("smoke_"):
                candidates.append(
                    candidate_from_df(
                        source_bucket="cluster_results_2026-03-31_15-10-05",
                        source_path=csv_path,
                        dataset="unknown",
                        method="unknown",
                        method_variant="smoke",
                        model="unknown",
                        shots=None,
                        baseline_type="other",
                        run_date_bucket="2026-03-31",
                        notes="smoke run",
                    )
                )
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="cluster_results_2026-03-31_15-10-05",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="nntp" if method == "nntp" else "llm",
                run_date_bucket="2026-03-31",
            )
        )
    return candidates


def collect_legacy_candidates() -> list[Candidate]:
    candidates: list[Candidate] = []

    for csv_path in sorted((OLDER_BUCKETS["legacy_gpt52_feb11"]).glob("*_eval_*.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_gpt52_feb11",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="llm",
                run_date_bucket="2026-02-11",
            )
        )

    for csv_path in sorted((OLDER_BUCKETS["legacy_mistral_feb16"]).glob("*_eval_*.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_mistral_feb16",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="llm",
                run_date_bucket="2026-02-16",
            )
        )

    for csv_path in sorted((OLDER_BUCKETS["legacy_gemini_mar02"]).glob("*_eval_*.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_gemini_mar02",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="llm",
                run_date_bucket="2026-03-02",
            )
        )

    for csv_path in sorted((OLDER_BUCKETS["legacy_children_jan21"]).glob("*.csv")):
        if not csv_path.name.startswith("children_handwritten_eval_"):
            continue
        m = re.fullmatch(r"children_handwritten_eval_(m[1-3])_(0shot|1shot)\.csv", csv_path.name)
        if not m:
            continue
        method = m.group(1)
        variant = m.group(2)
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_children_jan21",
                source_path=csv_path,
                dataset="children_handwritten",
                method=method,
                method_variant=variant,
                model="legacy_unspecified",
                shots=int(variant[0]),
                baseline_type="llm",
                run_date_bucket="2026-01-21",
            )
        )

    jan_root = OLDER_BUCKETS["legacy_jan_all_methods"]
    for subdir_name in ["0shot", "1shot"]:
        subdir = jan_root / subdir_name
        if not subdir.is_dir():
            continue
        shot = int(subdir_name[0])
        for csv_path in sorted(subdir.glob("*.csv")):
            name = csv_path.stem
            if name.startswith("summary_"):
                continue
            parsed = None
            if name.startswith("evaluation_qwen_"):
                continue
            m = re.fullmatch(r"evaluation_(m[1-3])_(.+)", name)
            if m:
                parsed = (normalize_dataset(m.group(2)), m.group(1), subdir_name, "legacy_unspecified", shot)
            m = re.fullmatch(r"(.+)_eval_qwen_(m[1-3])", name)
            if m:
                parsed = (normalize_dataset(m.group(1)), m.group(2), subdir_name, "qwen3-vl-8b-instruct", shot)
            m = re.fullmatch(r"(.+)_eval_(m[1-3])", name)
            if m and parsed is None:
                parsed = (normalize_dataset(m.group(1)), m.group(2), subdir_name, "legacy_unspecified", shot)
            if not parsed:
                continue
            dataset, method, variant, model, shot_value = parsed
            candidates.append(
                candidate_from_df(
                    source_bucket="legacy_jan_all_methods",
                    source_path=csv_path,
                    dataset=dataset,
                    method=method,
                    method_variant=variant,
                    model=model,
                    shots=shot_value,
                    baseline_type="llm",
                    run_date_bucket=f"2026-01-14_{subdir_name}",
                )
            )

    for csv_path in sorted((OLDER_BUCKETS["legacy_nntp_mar16"]).glob("*_eval_nntp.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_nntp_mar16",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="nntp",
                run_date_bucket="2026-03-16",
            )
        )

    for csv_path in sorted((OLDER_BUCKETS["legacy_nntp_mar20"]).glob("*_eval_nntp*.csv")):
        parsed = parse_cluster_eval_name(csv_path.name)
        if not parsed:
            continue
        dataset, method, variant, model, shots = parsed
        candidates.append(
            candidate_from_df(
                source_bucket="legacy_nntp_mar20",
                source_path=csv_path,
                dataset=dataset,
                method=method,
                method_variant=variant,
                model=model,
                shots=shots,
                baseline_type="nntp",
                run_date_bucket="2026-03-20",
            )
        )
    return candidates


def paper_bullinger_candidate() -> Candidate:
    line_counts = PAPER_BULLINGER["line_counts"]
    total_lines = line_counts["Subset1"] + line_counts["Subset2"]
    metrics = {}
    for metric in ["cer", "wer", "line_acc"]:
        s1 = PAPER_BULLINGER["Subset1"][metric]
        s2 = PAPER_BULLINGER["Subset2"][metric]
        metrics[metric] = (
            s1 * line_counts["Subset1"] + s2 * line_counts["Subset2"]
        ) / total_lines
    metrics["rev_line_acc"] = None
    metrics["exact_line_f1"] = None
    provider, model_family = model_info("NNTP (paper)")
    return Candidate(
        source_bucket="external_paper_2508.07904",
        source_csv=PAPER_URL,
        source_path=None,
        dataset="bullinger_handwritten",
        dataset_display=dataset_display("bullinger_handwritten"),
        dataset_group="handwritten_core",
        method="nntp",
        method_variant="paper_baseline",
        model="NNTP (paper)",
        model_family=model_family,
        provider=provider,
        shots=None,
        baseline_type="external_paper",
        ocr_status="external_paper",
        run_date_bucket="2025-08-paper",
        n_samples=69,
        metrics=metrics,
        notes=(
            "Paper baseline from arXiv:2508.07904 using PyLaia table values; overall metrics are "
            "approximated by line-count weighting over paper Subset 1/2 (20/49 letters, 902/1486 lines). "
            "Local Bullinger subset manifest contains 20/39 imported samples, so subset semantics align but counts differ."
        ),
    )


def precedence_rank(candidate: Candidate) -> int:
    if candidate.source_bucket == "thesis_bundle_2026-04-10":
        return 0
    if candidate.source_bucket == "thesis_run_root_2026-04-09":
        return 1
    if candidate.source_bucket == "cluster_results_2026-03-31_15-10-05":
        return 2
    return 3


def hard_exclusion_reason(candidate: Candidate) -> str | None:
    if candidate.method_variant == "smoke":
        return "smoke_run"
    if candidate.method_variant.startswith("0shot_prompt_") or "prompt_" in candidate.method_variant:
        return "prompt_ablation"
    if candidate.method == "m5" and candidate.method_variant == "separate_0shot":
        return "m5_separate_excluded"
    if candidate.dataset == "children_handwritten" and candidate.method in {"m2", "m3", "m4"}:
        if candidate.source_bucket == "cluster_results_2026-03-31_15-10-05":
            return "stale_children_ocr_pre_apr03"
    if candidate.dataset == "bullinger_handwritten" and candidate.method == "m5" and candidate.model == "mistral-large-2512":
        expected = EXPECTED_COUNTS["bullinger_handwritten"]
        if candidate.n_samples is None or candidate.n_samples != expected:
            return "incomplete_bullinger_m5_mistral"
    if candidate.dataset == "washington_handwritten" and candidate.method == "m4" and candidate.model == "gemini-3.1-pro-preview":
        if candidate.n_samples != EXPECTED_COUNTS["washington_handwritten"]:
            return "incomplete_washington_gemini_m4"
    return None


def qualifies_for_main(candidate: Candidate) -> bool:
    allowed_main_sources = {
        "thesis_bundle_2026-04-10",
        "thesis_run_root_2026-04-09",
        "cluster_results_2026-03-31_15-10-05",
        "external_paper_2508.07904",
    }
    if candidate.source_bucket not in allowed_main_sources:
        if not (
            candidate.source_bucket == "legacy_nntp_mar16"
            and candidate.dataset == "iam_handwritten_rwth_test_representative_20"
            and candidate.method == "nntp"
        ):
            return False
    if candidate.dataset not in CORE_DATASETS:
        return False
    if candidate.dataset == "bullinger_handwritten" and candidate.method == "nntp":
        return candidate.source_bucket == "external_paper_2508.07904"
    if candidate.method in {"m1", "m2", "m3", "m4"}:
        return candidate.method_variant == "0shot" and candidate.shots == 0 and candidate.model != "legacy_unspecified"
    if candidate.method == "m5":
        return candidate.method_variant == "context100"
    if candidate.method == "nntp":
        return candidate.method_variant in {"cv_macro", "legacy_single_run", "paper_baseline"}
    return False


def appendix_bucket(candidate: Candidate) -> bool:
    return candidate.source_bucket in {
        "legacy_jan_all_methods",
        "legacy_children_jan21",
        "legacy_gpt52_feb11",
        "legacy_mistral_feb16",
        "legacy_gemini_mar02",
        "legacy_nntp_mar16",
        "legacy_nntp_mar20",
    }


def select_rows(candidates: list[Candidate]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected_main: dict[tuple[str, str, str, str], Candidate] = {}
    audit: list[dict[str, Any]] = []

    exclusions = {candidate.source_csv: hard_exclusion_reason(candidate) for candidate in candidates}

    for candidate in candidates:
        if exclusions[candidate.source_csv]:
            continue
        if not qualifies_for_main(candidate):
            continue
        key = candidate.row_key
        chosen = selected_main.get(key)
        if chosen is None or precedence_rank(candidate) < precedence_rank(chosen):
            selected_main[key] = candidate

    selected_rows: list[dict[str, Any]] = []
    selected_sources = {candidate.source_csv: candidate for candidate in selected_main.values()}

    for candidate in sorted(selected_main.values(), key=lambda c: (c.dataset, c.method, c.model, c.method_variant)):
        row = candidate_to_row(candidate, scope="main", thesis_status="include_main")
        selected_rows.append(row)

    for candidate in candidates:
        reason = exclusions[candidate.source_csv]
        decision = "excluded"
        scope = ""
        notes = candidate.notes

        if candidate.source_csv in selected_sources:
            decision = "selected"
            scope = "main"
            reason = "selected_main_preferred_source"
        elif reason:
            decision = "excluded"
        elif appendix_bucket(candidate):
            if candidate.dataset == "children_handwritten" and candidate.method in {"m2", "m3", "m4", "m5"}:
                reason = "appendix_stale_or_superseded_ocr_sensitive"
                decision = "excluded"
            elif candidate.method == "nntp" and candidate.dataset == "iam_handwritten_rwth_test_representative_20":
                reason = "superseded_by_main_legacy_nntp"
                decision = "excluded"
            elif candidate.method == "nntp" and candidate.dataset == "bullinger_handwritten" and candidate.source_bucket == "legacy_nntp_mar16":
                row = candidate_to_row(
                    candidate,
                    scope="appendix",
                    thesis_status="include_appendix",
                    notes_suffix="Legacy local 10-sample exploratory Bullinger NNTP run.",
                )
                selected_rows.append(row)
                decision = "selected"
                scope = "appendix"
                reason = "selected_appendix_legacy_baseline"
            else:
                row = candidate_to_row(candidate, scope="appendix", thesis_status="include_appendix")
                selected_rows.append(row)
                decision = "selected"
                scope = "appendix"
                reason = "selected_appendix_historical"
        else:
            reason = reason or "superseded_or_non_target_source"

        audit.append(
            {
                "source_bucket": candidate.source_bucket,
                "source_csv": candidate.source_csv,
                "dataset": candidate.dataset,
                "method": candidate.method,
                "method_variant": candidate.method_variant,
                "model": candidate.model,
                "n_samples": candidate.n_samples,
                "decision": decision,
                "reason": reason,
                "scope": scope,
                "notes": notes,
            }
        )

    selected_rows = dedupe_selected_rows(selected_rows)
    return selected_rows, audit


def dedupe_selected_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    deduped: list[dict[str, Any]] = []
    for row in rows:
        key = (
            row["dataset"],
            row["method"],
            row["model"],
            row["method_variant"],
            row["subset_group"],
            row["source_csv"],
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def candidate_to_row(
    candidate: Candidate,
    *,
    scope: str,
    thesis_status: str,
    notes_suffix: str = "",
) -> dict[str, Any]:
    notes = candidate.notes.strip()
    if notes_suffix:
        notes = f"{notes} {notes_suffix}".strip()
    row = {
        "scope": scope,
        "dataset": candidate.dataset,
        "dataset_display": candidate.dataset_display,
        "dataset_group": candidate.dataset_group,
        "subset_group": "overall",
        "n_samples": candidate.n_samples,
        "method": candidate.method,
        "method_variant": candidate.method_variant,
        "model": candidate.model,
        "model_family": candidate.model_family,
        "provider": candidate.provider,
        "shots": candidate.shots,
        "baseline_type": candidate.baseline_type,
        "ocr_status": candidate.ocr_status,
        "source_bucket": candidate.source_bucket,
        "source_csv": candidate.source_csv,
        "run_date_bucket": candidate.run_date_bucket,
        "thesis_status": thesis_status,
        "notes": notes,
    }
    row.update(candidate.metrics)
    return row


def build_bullinger_subset_rows(master_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    subset_map = load_bullinger_subset_map()
    subset_rows: list[dict[str, Any]] = []

    for row in master_rows:
        if row["dataset"] != "bullinger_handwritten":
            continue

        if row["source_bucket"] == "external_paper_2508.07904":
            subset_rows.extend(bullinger_paper_subset_rows(row))
            continue

        source_csv = row["source_csv"]
        if not source_csv.startswith("/"):
            continue
        df = read_eval_df(Path(source_csv))
        samples = sample_rows(df, "bullinger_handwritten")
        if samples.empty:
            continue
        for subset_name in ["Subset1", "Subset2"]:
            subset_samples = samples[samples["id"].astype(str).map(subset_map.get) == subset_name]
            if subset_samples.empty:
                continue
            subset_rows.append(
                make_subset_row(
                    row,
                    subset_name=subset_name,
                    subset_label="historian" if subset_name == "Subset1" else "diplomatic",
                    n_samples=int(len(subset_samples)),
                    metrics={metric: float_or_none(subset_samples[metric].mean()) for metric in METRIC_COLUMNS},
                    notes=row["notes"],
                )
            )
        subset_rows.append(
            make_subset_row(
                row,
                subset_name="overall",
                subset_label="overall",
                n_samples=int(len(samples)),
                metrics={metric: row[metric] for metric in METRIC_COLUMNS},
                notes=row["notes"],
            )
        )

    return subset_rows


def bullinger_paper_subset_rows(row: dict[str, Any]) -> list[dict[str, Any]]:
    subset_rows: list[dict[str, Any]] = []
    paper_notes = (
        f"{row['notes']} Subset values are copied from Tables 3 and 4 for the PyLaia row in {PAPER_URL}. "
        "Subset metrics come from the paper's 20/49-letter corrected split and therefore are not directly count-matched "
        "to the local 20/39 imported Bullinger manifest."
    )
    subset_rows.append(
        make_subset_row(
            row,
            subset_name="Subset1",
            subset_label="historian",
            n_samples=20,
            metrics={
                "wer": PAPER_BULLINGER["Subset1"]["wer"],
                "cer": PAPER_BULLINGER["Subset1"]["cer"],
                "line_acc": PAPER_BULLINGER["Subset1"]["line_acc"],
                "rev_line_acc": None,
                "exact_line_f1": None,
            },
            notes=paper_notes,
        )
    )
    subset_rows.append(
        make_subset_row(
            row,
            subset_name="Subset2",
            subset_label="diplomatic",
            n_samples=49,
            metrics={
                "wer": PAPER_BULLINGER["Subset2"]["wer"],
                "cer": PAPER_BULLINGER["Subset2"]["cer"],
                "line_acc": PAPER_BULLINGER["Subset2"]["line_acc"],
                "rev_line_acc": None,
                "exact_line_f1": None,
            },
            notes=paper_notes,
        )
    )
    subset_rows.append(
        make_subset_row(
            row,
            subset_name="overall",
            subset_label="overall",
            n_samples=69,
            metrics={metric: row.get(metric) for metric in METRIC_COLUMNS},
            notes=f"{paper_notes} Overall metrics are approximated by line-count weighting using Table 1 line counts (902/1486).",
        )
    )
    return subset_rows


def make_subset_row(
    master_row: dict[str, Any],
    *,
    subset_name: str,
    subset_label: str,
    n_samples: int,
    metrics: dict[str, float | None],
    notes: str,
) -> dict[str, Any]:
    row = {
        "scope": master_row["scope"],
        "dataset": master_row["dataset"],
        "dataset_display": master_row["dataset_display"],
        "subset_group": subset_name,
        "subset_label": subset_label,
        "n_samples": n_samples,
        "method": master_row["method"],
        "method_variant": master_row["method_variant"],
        "model": master_row["model"],
        "baseline_type": master_row["baseline_type"],
        "source_bucket": master_row["source_bucket"],
        "source_csv": master_row["source_csv"],
        "thesis_status": master_row["thesis_status"],
        "notes": notes,
    }
    row.update(metrics)
    return row


def build_main_and_appendix_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = []
    candidates.extend(collect_bundle_candidates())
    candidates.extend(collect_runroot_candidates())
    candidates.extend(collect_cluster_results_candidates())
    candidates.extend(collect_legacy_candidates())
    candidates.append(paper_bullinger_candidate())
    return select_rows(candidates)


def sort_master_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    method_order = {"nntp": 0, "m1": 1, "m2": 2, "m3": 3, "m4": 4, "m5": 5}
    scope_order = {"main": 0, "appendix": 1}
    return sorted(
        rows,
        key=lambda row: (
            scope_order.get(row["scope"], 9),
            row["dataset_group"],
            row["dataset"],
            method_order.get(row["method"], 9),
            row["model"],
            row["method_variant"],
        ),
    )


def autosize_sheet(worksheet: Any, dataframe: pd.DataFrame) -> None:
    for idx, column in enumerate(dataframe.columns):
        max_len = max(
            [len(str(column))]
            + [len(str(value)) for value in dataframe[column].fillna("").tolist()]
        )
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max(max_len + 2, 12), 60)


def compare_label(delta: float | None) -> str:
    if delta is None:
        return ""
    if math.isclose(delta, 0.0, abs_tol=1e-12):
        return "ties NNTP"
    if delta > 0:
        return "beats NNTP"
    return "below NNTP"


def baseline_provenance_label(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if row["baseline_type"] == "external_paper":
        return "paper"
    return str(row["method_variant"])


def build_summary_tables(
    master_rows: list[dict[str, Any]],
    subset_rows: list[dict[str, Any]],
) -> list[tuple[str, pd.DataFrame]]:
    master_df = pd.DataFrame(master_rows, columns=MASTER_COLUMNS)
    subset_df = pd.DataFrame(subset_rows, columns=SUBSET_COLUMNS)
    core_main = master_df[
        (master_df["scope"] == "main") & (master_df["dataset_group"] == "handwritten_core")
    ].copy()
    core_llm = core_main[core_main["baseline_type"] == "llm"].copy()

    dataset_rows: list[dict[str, Any]] = []
    for dataset in DATASET_ORDER:
        dataset_llm = core_llm[core_llm["dataset"] == dataset]
        if dataset_llm.empty:
            continue

        best_rows_by_method: dict[str, pd.Series | None] = {}
        for method in METHOD_ORDER:
            method_rows = dataset_llm[dataset_llm["method"] == method].sort_values(
                ["line_acc", "model"], ascending=[False, True]
            )
            best_rows_by_method[method] = None if method_rows.empty else method_rows.iloc[0]

        overall_best_candidates = [row for row in best_rows_by_method.values() if row is not None]
        if not overall_best_candidates:
            continue
        overall_best = sorted(
            overall_best_candidates,
            key=lambda row: (-float(row["line_acc"]), str(row["method"]), str(row["model"])),
        )[0]

        nntp_rows = core_main[(core_main["dataset"] == dataset) & (core_main["method"] == "nntp")].sort_values(
            ["line_acc", "model"], ascending=[False, True]
        )
        baseline = None if nntp_rows.empty else nntp_rows.iloc[0]
        baseline_line_acc = None if baseline is None else baseline["line_acc"]
        delta = None if baseline is None else float(overall_best["line_acc"] - baseline["line_acc"])

        dataset_rows.append(
            {
                "dataset": overall_best["dataset_display"],
                "n_samples": int(overall_best["n_samples"]),
                "nntp_line_acc": baseline_line_acc,
                "nntp_provenance": baseline_provenance_label(baseline),
                "m1_max_line_acc": None if best_rows_by_method["m1"] is None else best_rows_by_method["m1"]["line_acc"],
                "m2_max_line_acc": None if best_rows_by_method["m2"] is None else best_rows_by_method["m2"]["line_acc"],
                "m3_max_line_acc": None if best_rows_by_method["m3"] is None else best_rows_by_method["m3"]["line_acc"],
                "m4_max_line_acc": None if best_rows_by_method["m4"] is None else best_rows_by_method["m4"]["line_acc"],
                "m5_max_line_acc": None if best_rows_by_method["m5"] is None else best_rows_by_method["m5"]["line_acc"],
                "best_method": overall_best["method"],
                "best_model": overall_best["model"],
                "best_line_acc": overall_best["line_acc"],
                "delta_best_vs_nntp": delta,
            }
        )

    dataset_summary_df = pd.DataFrame(dataset_rows, columns=DATASET_SUMMARY_COLUMNS)

    method_summary_rows: list[dict[str, Any]] = []
    for method in METHOD_ORDER:
        method_df = core_llm[core_llm["method"] == method].copy()
        if method_df.empty:
            continue
        grouped = (
            method_df.groupby("model", dropna=False)
            .agg(
                avg_line_acc=("line_acc", "mean"),
                datasets_covered=("dataset", "nunique"),
            )
            .reset_index()
            .sort_values(["datasets_covered", "avg_line_acc", "model"], ascending=[False, False, True])
        )
        best_model = grouped.iloc[0]["model"]
        best_rows = method_df[method_df["model"] == best_model].sort_values(
            ["line_acc", "dataset"], ascending=[False, True]
        )
        best_dataset_row = best_rows.iloc[0]
        method_summary_rows.append(
            {
                "method": method,
                "best_model": best_model,
                "mean_line_acc": grouped.iloc[0]["avg_line_acc"],
                "datasets_covered": int(grouped.iloc[0]["datasets_covered"]),
                "best_dataset": best_dataset_row["dataset_display"],
                "best_dataset_line_acc": best_dataset_row["line_acc"],
            }
        )
    method_summary_df = pd.DataFrame(method_summary_rows, columns=METHOD_SUMMARY_COLUMNS)

    bullinger_rows: list[dict[str, Any]] = []
    bullinger_main_subsets = subset_df[
        (subset_df["scope"] == "main") & (subset_df["dataset"] == "bullinger_handwritten")
    ].copy()
    for subset_name in ["Subset1", "Subset2", "overall"]:
        rows = bullinger_main_subsets[bullinger_main_subsets["subset_group"] == subset_name]
        paper_rows = rows[rows["baseline_type"] == "external_paper"].sort_values(
            ["line_acc", "model"], ascending=[False, True]
        )
        best_rows_by_method: dict[str, pd.Series | None] = {}
        for method in METHOD_ORDER:
            method_rows = rows[rows["method"] == method].sort_values(
                ["line_acc", "model"], ascending=[False, True]
            )
            best_rows_by_method[method] = None if method_rows.empty else method_rows.iloc[0]

        overall_best_candidates = [row for row in best_rows_by_method.values() if row is not None]
        if not overall_best_candidates:
            continue
        best = sorted(
            overall_best_candidates,
            key=lambda row: (-float(row["line_acc"]), str(row["method"]), str(row["model"])),
        )[0]
        paper = paper_rows.iloc[0] if not paper_rows.empty else None
        delta = float(best["line_acc"] - paper["line_acc"]) if paper is not None else None
        bullinger_rows.append(
            {
                "subset": f"{subset_name} ({best['subset_label']})",
                "local_n_samples": int(best["n_samples"]),
                "nntp_paper_line_acc": None if paper is None else paper["line_acc"],
                "m1_max_line_acc": None if best_rows_by_method["m1"] is None else best_rows_by_method["m1"]["line_acc"],
                "m2_max_line_acc": None if best_rows_by_method["m2"] is None else best_rows_by_method["m2"]["line_acc"],
                "m3_max_line_acc": None if best_rows_by_method["m3"] is None else best_rows_by_method["m3"]["line_acc"],
                "m4_max_line_acc": None if best_rows_by_method["m4"] is None else best_rows_by_method["m4"]["line_acc"],
                "m5_max_line_acc": None if best_rows_by_method["m5"] is None else best_rows_by_method["m5"]["line_acc"],
                "best_method": best["method"],
                "best_model": best["model"],
                "best_line_acc": best["line_acc"],
                "delta_best_vs_nntp": delta,
            }
        )
    bullinger_summary_df = pd.DataFrame(bullinger_rows, columns=BULLINGER_SUMMARY_COLUMNS)

    return [
        ("Maximum line_acc by dataset and method", dataset_summary_df),
        ("Best model per method (mean line_acc across handwritten datasets)", method_summary_df),
        ("Bullinger subset comparison against NNTP paper baseline", bullinger_summary_df),
    ]


def write_table_to_sheet(worksheet: Any, dataframe: pd.DataFrame, start_row: int) -> int:
    header_font = Font(bold=True)
    for col_idx, column in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=start_row, column=col_idx, value=column)
        cell.font = header_font

    row_idx = start_row + 1
    for row in dataframe.itertuples(index=False):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, float):
                cell.number_format = "0.0000"
        row_idx += 1
    return row_idx - 1


def write_summary_sheet(workbook: Any, summary_tables: list[tuple[str, pd.DataFrame]]) -> None:
    worksheet = workbook.create_sheet(SUMMARY_SHEET_NAME, 0)
    title_font = Font(bold=True)
    current_row = 1

    first = True
    for title, dataframe in summary_tables:
        if dataframe.empty:
            continue
        if first:
            current_row = write_table_to_sheet(worksheet, dataframe, current_row)
            first = False
        else:
            current_row += 2
            title_cell = worksheet.cell(row=current_row, column=1, value=title)
            title_cell.font = title_font
            current_row += 1
            current_row = write_table_to_sheet(worksheet, dataframe, current_row)

    note_row = current_row + 2
    note_cell = worksheet.cell(
        row=note_row,
        column=1,
        value=(
            "Note: Bullinger NNTP uses the external paper baseline from arXiv:2508.07904; "
            "the paper subset counts differ from the local imported sample counts."
        ),
    )
    note_cell.font = title_font
    worksheet.freeze_panes = "A2"
    autosize_sheet(worksheet, pd.DataFrame(worksheet.values))


def write_workbook(
    master_rows: list[dict[str, Any]],
    subset_rows: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
    out_path: Path,
) -> None:
    master_df = pd.DataFrame(sort_master_rows(master_rows), columns=MASTER_COLUMNS)
    subset_df = pd.DataFrame(subset_rows, columns=SUBSET_COLUMNS)
    audit_df = pd.DataFrame(audit_rows, columns=AUDIT_COLUMNS).sort_values(
        ["decision", "source_bucket", "dataset", "method", "model", "source_csv"]
    )
    summary_tables = build_summary_tables(master_rows, subset_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="results_master", index=False)
        subset_df.to_excel(writer, sheet_name="bullinger_subsets", index=False)
        audit_df.to_excel(writer, sheet_name="source_audit", index=False)

        workbook = writer.book
        write_summary_sheet(workbook, summary_tables)
        for sheet_name, dataframe in [
            ("results_master", master_df),
            ("bullinger_subsets", subset_df),
            ("source_audit", audit_df),
        ]:
            worksheet = workbook[sheet_name]
            worksheet.freeze_panes = "A2"
            autosize_sheet(worksheet, dataframe)


def main() -> None:
    args = parse_args()
    out_path = Path(args.out).expanduser().resolve()
    master_rows, audit_rows = build_main_and_appendix_rows()
    subset_rows = build_bullinger_subset_rows(master_rows)
    write_workbook(master_rows, subset_rows, audit_rows, out_path)
    print(f"Wrote workbook to {out_path}")
    print(f"results_master rows: {len(master_rows)}")
    print(f"bullinger_subsets rows: {len(subset_rows)}")
    print(f"source_audit rows: {len(audit_rows)}")


if __name__ == "__main__":
    main()
