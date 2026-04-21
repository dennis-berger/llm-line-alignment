#!/usr/bin/env python3
"""Assemble the final immutable thesis results release folder."""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REPO_ROOT = Path("/Users/dennisberger/Documents/Privat/llm-line-alignment")
PYTHON_BIN = REPO_ROOT / ".venv" / "bin" / "python"
THESIS_PREDICTIONS_ROOT = Path(
    "/Users/dennisberger/Library/Mobile Documents/com~apple~CloudDocs/Dokumente/Uni/Master_Thesis/predictions"
)
THESIS_BUNDLE = THESIS_PREDICTIONS_ROOT / "handwritten_thesis_bundle_2026-04-10"
THESIS_RUN_ROOT = REPO_ROOT / "cluster_runs" / "thesis_handwritten_completion_2026-04-09"
QWEN_RUN_ROOT = REPO_ROOT / "cluster_runs" / "handwritten_m5_context100_qwen3_32b_2026-04-17"
CLUSTER_RESULTS_MAR31 = THESIS_PREDICTIONS_ROOT / "cluster_results_2026-03-31_15-10-05"
DEFAULT_OUT_DIR = THESIS_PREDICTIONS_ROOT / "thesis_results_final_2026-04-20"
WORKBOOK_NAME = "handwritten_thesis_results_master.xlsx"
ALLOWED_PARTIAL_M5 = "bullinger_handwritten:qwen3-vl-32b-instruct"
HANDWRITTEN_RESULTS_DIRNAME = "handwritten_results"
SOURCES_DIRNAME = "sources"
BUILDER_SCRIPT = REPO_ROOT / "scripts" / "build_handwritten_thesis_results_master.py"


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR), help="Final release directory.")
    ap.add_argument(
        "--thesis-predictions-root",
        default=str(THESIS_PREDICTIONS_ROOT),
        help="Predictions root containing the historical thesis folders.",
    )
    ap.add_argument("--thesis-bundle", default=str(THESIS_BUNDLE), help="Primary handwritten bundle.")
    ap.add_argument("--thesis-run-root", default=str(THESIS_RUN_ROOT), help="Primary handwritten rerun root.")
    ap.add_argument("--qwen-run-root", default=str(QWEN_RUN_ROOT), help="Synced Qwen handwritten run root.")
    ap.add_argument(
        "--cluster-results-mar31",
        default=str(CLUSTER_RESULTS_MAR31),
        help="March 31 cluster results root.",
    )
    ap.add_argument(
        "--allow-partial-m5",
        action="append",
        default=[ALLOWED_PARTIAL_M5],
        help="Allow a partial M5 row to remain in the release workbook.",
    )
    ap.add_argument("--overwrite", action="store_true", help="Replace the output directory if it exists.")
    return ap.parse_args()


def ensure_clean_dir(path: Path, overwrite: bool) -> None:
    if path.exists():
        if not overwrite:
            raise FileExistsError(f"Release directory already exists: {path}")
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def is_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def method_dir_name(row: pd.Series) -> str:
    method = str(row["method"])
    method_variant = str(row["method_variant"])
    if method == "m5" and method_variant == "context100":
        return "m5_context100"
    return method


def copy_any(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    if src.is_dir():
        shutil.copytree(src, dst, dirs_exist_ok=True)
    else:
        shutil.copy2(src, dst)


def copy_dir_contents(src_dir: Path, dst_dir: Path) -> None:
    dst_dir.mkdir(parents=True, exist_ok=True)
    for child in src_dir.iterdir():
        copy_any(child, dst_dir / child.name)


def artifact_root_for_csv(csv_path: Path) -> Path | None:
    if csv_path.name in {"evaluation.csv", "eval.csv", "eval_m5_context100.csv"}:
        return csv_path.parent
    return None


def candidate_prediction_paths(csv_path: Path) -> list[Path]:
    candidates: list[Path] = []
    name = csv_path.name
    stem = csv_path.stem
    parent = csv_path.parent

    if name in {"evaluation.csv", "eval.csv", "eval_m5_context100.csv"}:
        for child in ["predictions", "predictions_m5_context100"]:
            path = parent / child
            if path.is_dir():
                candidates.append(path)

    if "_eval_" in stem:
        pred_dir = parent / stem.replace("_eval_", "_predictions_")
        if pred_dir.is_dir():
            candidates.append(pred_dir)

    if name.endswith("_eval_nntp_cv.csv"):
        pred_dir = parent / "predictions"
        if pred_dir.is_dir():
            candidates.append(pred_dir)

    if name.endswith("_eval_nntp_cv_macro.csv"):
        for suffix in ["train_a_test_b", "train_b_test_a"]:
            pred_dir = parent / name.replace("_eval_nntp_cv_macro.csv", f"_predictions_nntp_{suffix}")
            if pred_dir.is_dir():
                candidates.append(pred_dir)

    if name.endswith("_eval_nntp_train_a_test_b.csv") or name.endswith("_eval_nntp_train_b_test_a.csv"):
        pred_dir = parent / stem.replace("_eval_", "_predictions_")
        if pred_dir.is_dir():
            candidates.append(pred_dir)

    if name.endswith("_eval_nntp_iam_zero_adapt.csv"):
        pred_dir = parent / stem.replace("_eval_", "_predictions_")
        if pred_dir.is_dir():
            candidates.append(pred_dir)

    deduped: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        if candidate not in seen:
            seen.add(candidate)
            deduped.append(candidate)
    return deduped


def count_files(root: Path | None) -> int:
    if root is None or not root.exists():
        return 0
    return sum(1 for path in root.rglob("*") if path.is_file())


def copy_predictions_to_standard_root(source_dirs: list[Path], dest_predictions: Path) -> None:
    if not source_dirs:
        return
    if len(source_dirs) == 1:
        source = source_dirs[0]
        if source.name == "predictions":
            copy_any(source, dest_predictions)
            return
        if any(child.is_dir() for child in source.iterdir()):
            copy_any(source, dest_predictions / source.name)
            return
        copy_dir_contents(source, dest_predictions)
        return

    for source in source_dirs:
        copy_any(source, dest_predictions / source.name)


def stage_run_artifacts(source_csv: Path, dest_run_dir: Path) -> dict[str, int]:
    dest_run_dir.mkdir(parents=True, exist_ok=True)
    artifact_root = artifact_root_for_csv(source_csv)

    shutil.copy2(source_csv, dest_run_dir / "evaluation.csv")

    copied_predictions = False
    if artifact_root is not None:
        for child_name in ["predictions", "predictions_m5_context100", "traces", "checkpoints"]:
            child = artifact_root / child_name
            if not child.exists():
                continue
            target_name = "predictions" if child_name == "predictions_m5_context100" else child_name
            copy_any(child, dest_run_dir / target_name)
            if target_name == "predictions":
                copied_predictions = True

    if not copied_predictions:
        prediction_dirs = candidate_prediction_paths(source_csv)
        copy_predictions_to_standard_root(prediction_dirs, dest_run_dir / "predictions")

    predictions_dir = dest_run_dir / "predictions"
    traces_dir = dest_run_dir / "traces"
    checkpoints_dir = dest_run_dir / "checkpoints"
    return {
        "predictions_files": count_files(predictions_dir),
        "trace_files": count_files(traces_dir),
        "checkpoint_files": count_files(checkpoints_dir),
    }


def path_to_sources_rel(path: Path, predictions_root: Path) -> Path:
    resolved = path.expanduser().resolve()
    if resolved.is_relative_to(predictions_root):
        return Path(SOURCES_DIRNAME) / "thesis_predictions" / resolved.relative_to(predictions_root)
    if resolved.is_relative_to(REPO_ROOT):
        return Path(SOURCES_DIRNAME) / "repo" / resolved.relative_to(REPO_ROOT)
    return Path(SOURCES_DIRNAME) / "external_local" / resolved.as_posix().lstrip("/")


def gather_workbook_frames(workbook_path: Path) -> tuple[list[str], dict[str, pd.DataFrame]]:
    xls = pd.ExcelFile(workbook_path)
    sheet_names = xls.sheet_names
    frames = {sheet: pd.read_excel(workbook_path, sheet_name=sheet) for sheet in sheet_names}
    return sheet_names, frames


def gather_source_values(frames: dict[str, pd.DataFrame]) -> list[str]:
    sources: list[str] = []
    for frame in frames.values():
        for column in frame.columns:
            if "source_csv" not in str(column):
                continue
            for value in frame[column].dropna().astype(str):
                if value and value not in sources:
                    sources.append(value)
    return sources


def build_builder_command(args: argparse.Namespace, temp_workbook: Path) -> list[str]:
    cmd = [
        str(PYTHON_BIN if PYTHON_BIN.is_file() else "python3"),
        str(BUILDER_SCRIPT),
        "--out",
        str(temp_workbook),
        "--thesis-predictions-root",
        str(Path(args.thesis_predictions_root).expanduser().resolve()),
        "--thesis-bundle",
        str(Path(args.thesis_bundle).expanduser().resolve()),
        "--thesis-run-root",
        str(Path(args.thesis_run_root).expanduser().resolve()),
        "--qwen-run-root",
        str(Path(args.qwen_run_root).expanduser().resolve()),
        "--cluster-results-mar31",
        str(Path(args.cluster_results_mar31).expanduser().resolve()),
    ]
    for item in args.allow_partial_m5:
        cmd.extend(["--allow-partial-m5", item])
    return cmd


def stage_handwritten_rows(
    master_df: pd.DataFrame,
    final_dir: Path,
) -> tuple[dict[str, str], list[dict[str, Any]], list[dict[str, Any]]]:
    handwritten_root = final_dir / HANDWRITTEN_RESULTS_DIRNAME
    staged_source_map: dict[str, str] = {}
    manifest_rows: list[dict[str, Any]] = []
    partial_rows: list[dict[str, Any]] = []

    main_rows = master_df[
        (master_df["scope"] == "main")
        & (master_df["dataset_group"] == "handwritten_core")
    ].copy()

    for row in main_rows.itertuples(index=False):
        source_csv = str(row.source_csv)
        if not source_csv or is_url(source_csv):
            continue
        source_path = Path(source_csv).expanduser().resolve()
        dest_run_dir = handwritten_root / str(row.dataset) / method_dir_name(pd.Series(row._asdict())) / str(row.model)
        counts = stage_run_artifacts(source_path, dest_run_dir)
        relative_csv = (dest_run_dir / "evaluation.csv").relative_to(final_dir).as_posix()
        staged_source_map[source_csv] = relative_csv
        manifest_row = {
            "kind": "handwritten_run",
            "dataset": str(row.dataset),
            "method": str(row.method),
            "method_variant": str(row.method_variant),
            "model": str(row.model),
            "source_csv": source_csv,
            "dest_csv": relative_csv,
            "thesis_status": str(row.thesis_status),
            "notes": "" if pd.isna(row.notes) else str(row.notes),
            **counts,
        }
        manifest_rows.append(manifest_row)
        if str(row.thesis_status) == "partial":
            partial_rows.append(manifest_row)

    return staged_source_map, manifest_rows, partial_rows


def copy_support_sources(
    final_dir: Path,
    source_values: list[str],
    *,
    source_map: dict[str, str],
    predictions_root: Path,
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    copied_map = dict(source_map)
    copied_items: list[dict[str, Any]] = []
    copied_paths: set[str] = set()

    for source in source_values:
        if source in copied_map:
            continue
        if is_url(source):
            copied_map[source] = source
            copied_items.append({"kind": "external_url", "source": source, "relative_path": source})
            continue

        src_path = Path(source).expanduser().resolve()
        if not src_path.exists():
            copied_items.append({"kind": "missing", "source": source, "relative_path": ""})
            continue

        target_rel = path_to_sources_rel(src_path, predictions_root)
        copy_any(src_path, final_dir / target_rel)
        copied_map[source] = target_rel.as_posix()
        copied_items.append({"kind": "csv", "source": source, "relative_path": target_rel.as_posix()})

        for pred_dir in candidate_prediction_paths(src_path):
            pred_key = str(pred_dir.resolve())
            if pred_key in copied_paths:
                continue
            copied_paths.add(pred_key)
            pred_rel = path_to_sources_rel(pred_dir.resolve(), predictions_root)
            copy_any(pred_dir.resolve(), final_dir / pred_rel)
            copied_items.append(
                {
                    "kind": "predictions_dir",
                    "source": pred_key,
                    "relative_path": pred_rel.as_posix(),
                }
            )

    return copied_map, copied_items


def rewrite_workbook_source_paths(source_workbook: Path, out_workbook: Path, source_map: dict[str, str]) -> None:
    shutil.copy2(source_workbook, out_workbook)
    workbook = load_workbook(out_workbook)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in source_map:
                    cell.value = source_map[cell.value]
    workbook.save(out_workbook)


def write_release_manifest(
    final_dir: Path,
    *,
    release_name: str,
    builder_command: list[str],
    staged_rows: list[dict[str, Any]],
    support_items: list[dict[str, Any]],
    partial_rows: list[dict[str, Any]],
    audit_df: pd.DataFrame,
) -> None:
    excluded_rows = audit_df[audit_df["decision"] == "excluded"].copy()
    selected_rows = audit_df[audit_df["decision"].isin(["selected", "selected_partial"])].copy()
    manifest = {
        "release_name": release_name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "selection_rule": (
            "For each dataset/method/model combination, use the newest authoritative result. "
            "Prefer complete runs over older runs; Bullinger handwritten Qwen m5 context100 is "
            "included as an explicitly partial exception and excluded from summary maxima."
        ),
        "builder_command": builder_command,
        "included_artifacts": staged_rows + support_items,
        "partial_artifacts": partial_rows,
        "selected_rows": selected_rows.to_dict(orient="records"),
        "excluded_superseded_artifacts": excluded_rows.to_dict(orient="records"),
    }
    (final_dir / "release_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def write_readme(final_dir: Path) -> None:
    readme = """# Final Thesis Results Release

This folder is the single thesis reference root for the final handwritten results state.

## Layout

- `handwritten_results/`: source of truth for the handwritten raw artifacts used in the thesis workbook.
- `sources/`: appendix and legacy workbook sources that are still referenced but are not part of the main handwritten raw tree.
- `handwritten_thesis_results_master.xlsx`: final workbook rebuilt for this release.
- `release_manifest.json`: provenance, inclusion, partial-status, and excluded-source manifest.

## Notes

- Bullinger handwritten `m5 context100` with `qwen3-vl-32b-instruct` is included as a partial run only.
- The partial Bullinger Qwen row remains visible in the workbook, but it is excluded from technical-summary maxima and best-model comparisons.
- Older folders under the thesis `predictions/` root are historical intermediates and are superseded by this release folder.
"""
    (final_dir / "README.md").write_text(readme, encoding="utf-8")


def main() -> None:
    args = parse_args()
    final_dir = Path(args.out_dir).expanduser().resolve()
    ensure_clean_dir(final_dir, args.overwrite)

    predictions_root = Path(args.thesis_predictions_root).expanduser().resolve()
    builder_command: list[str]

    with tempfile.TemporaryDirectory(prefix="final_thesis_release_") as tmp_dir_name:
        tmp_dir = Path(tmp_dir_name)
        temp_workbook = tmp_dir / WORKBOOK_NAME
        builder_command = build_builder_command(args, temp_workbook)
        subprocess.run(builder_command, cwd=REPO_ROOT, check=True)

        sheet_names, frames = gather_workbook_frames(temp_workbook)
        master_df = frames["results_master"].copy()
        audit_df = frames["source_audit"].copy()
        source_values = gather_source_values(frames)

        staged_source_map, staged_rows, partial_rows = stage_handwritten_rows(master_df, final_dir)
        source_map, support_items = copy_support_sources(
            final_dir,
            source_values,
            source_map=staged_source_map,
            predictions_root=predictions_root,
        )

        final_workbook = final_dir / WORKBOOK_NAME
        rewrite_workbook_source_paths(temp_workbook, final_workbook, source_map)

    write_release_manifest(
        final_dir,
        release_name=final_dir.name,
        builder_command=builder_command,
        staged_rows=staged_rows,
        support_items=support_items,
        partial_rows=partial_rows,
        audit_df=audit_df,
    )
    write_readme(final_dir)

    print(f"Created final release folder at {final_dir}")
    print(f"Workbook: {final_dir / WORKBOOK_NAME}")
    print(f"Handwritten staged rows: {len(staged_rows)}")
    print(f"Support items copied: {len(support_items)}")
    print(f"Partial rows: {len(partial_rows)}")


if __name__ == "__main__":
    main()
