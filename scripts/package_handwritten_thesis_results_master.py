#!/usr/bin/env python3
"""Package the thesis master workbook with relative-path sources and predictions."""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


THESIS_PREDICTIONS_ROOT = Path(
    "/Users/dennisberger/Library/Mobile Documents/com~apple~CloudDocs/Dokumente/Uni/Master_Thesis/predictions"
)
DEFAULT_WORKBOOK = THESIS_PREDICTIONS_ROOT / "handwritten_thesis_results_master.xlsx"
REPO_ROOT = Path("/Users/dennisberger/Documents/Privat/llm-line-alignment")
DEFAULT_OUT_DIR = THESIS_PREDICTIONS_ROOT / "handwritten_thesis_results_master_package_2026-04-10"
WORKBOOK_NAME = "handwritten_thesis_results_master.xlsx"


@dataclass(frozen=True)
class CopyResult:
    source: str
    target_rel: str


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Workbook to package.")
    ap.add_argument("--out-dir", default=str(DEFAULT_OUT_DIR), help="Output package directory.")
    ap.add_argument("--overwrite", action="store_true", help="Replace the output directory if it exists.")
    return ap.parse_args()


def ensure_clean_dir(path: Path, overwrite: bool) -> None:
    if path.exists():
        if not overwrite:
            raise FileExistsError(f"Package directory already exists: {path}")
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def path_to_rel_copy(path: Path) -> Path:
    resolved = path.expanduser().resolve()
    if resolved.is_relative_to(THESIS_PREDICTIONS_ROOT):
        return Path("sources") / "thesis_predictions" / resolved.relative_to(THESIS_PREDICTIONS_ROOT)
    if resolved.is_relative_to(REPO_ROOT):
        return Path("sources") / "repo" / resolved.relative_to(REPO_ROOT)
    return Path("sources") / "external_local" / resolved.as_posix().lstrip("/")


def copy_any(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    if src.is_dir():
        shutil.copytree(src, dst, dirs_exist_ok=True)
    else:
        shutil.copy2(src, dst)


def is_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def rewrite_source_value(source_csv: str, copied_files: dict[str, str]) -> str:
    if not source_csv or is_url(source_csv):
        return source_csv
    return copied_files.get(source_csv, source_csv)


def candidate_prediction_paths(csv_path: Path) -> list[Path]:
    candidates: list[Path] = []
    name = csv_path.name
    stem = csv_path.stem
    parent = csv_path.parent

    if name in {"evaluation.csv", "eval.csv", "eval_m5_context100.csv"}:
        for child in ["predictions", "predictions_m5_context100"]:
            p = parent / child
            if p.is_dir():
                candidates.append(p)

    if "_eval_" in stem:
        pred_dir = parent / stem.replace("_eval_", "_predictions_")
        if pred_dir.is_dir():
            candidates.append(pred_dir)

    if name.endswith("_eval_nntp_cv.csv"):
        p = parent / "predictions"
        if p.is_dir():
            candidates.append(p)

    if name.endswith("_eval_nntp_cv_macro.csv"):
        for suffix in ["train_a_test_b", "train_b_test_a"]:
            p = parent / name.replace("_eval_nntp_cv_macro.csv", f"_predictions_nntp_{suffix}")
            if p.is_dir():
                candidates.append(p)

    if name.endswith("_eval_nntp_train_a_test_b.csv") or name.endswith("_eval_nntp_train_b_test_a.csv"):
        p = parent / stem.replace("_eval_", "_predictions_")
        if p.is_dir():
            candidates.append(p)

    if name.endswith("_eval_nntp_iam_zero_adapt.csv"):
        p = parent / stem.replace("_eval_", "_predictions_")
        if p.is_dir():
            candidates.append(p)

    deduped: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        if candidate not in seen:
            seen.add(candidate)
            deduped.append(candidate)
    return deduped


def gather_source_values(workbook: Path) -> tuple[list[str], dict[str, pd.DataFrame], list[str]]:
    xls = pd.ExcelFile(workbook)
    sheet_names = xls.sheet_names
    frames = {sheet: pd.read_excel(workbook, sheet_name=sheet) for sheet in sheet_names}
    sources: list[str] = []
    for frame in frames.values():
        for column in frame.columns:
            if "source_csv" not in str(column):
                continue
            for value in frame[column].dropna().astype(str):
                if value and value not in sources:
                    sources.append(value)
    return sheet_names, frames, sources


def package_sources(package_dir: Path, source_values: list[str]) -> tuple[dict[str, str], list[dict[str, Any]]]:
    copied_files: dict[str, str] = {}
    manifest_rows: list[dict[str, Any]] = []
    copied_prediction_dirs: set[str] = set()

    for source in source_values:
        if is_url(source):
            manifest_rows.append({"source": source, "relative_path": source, "kind": "external_url"})
            continue

        src_path = Path(source).expanduser().resolve()
        if not src_path.exists():
            manifest_rows.append({"source": source, "relative_path": "", "kind": "missing"})
            continue

        target_rel = path_to_rel_copy(src_path)
        copy_any(src_path, package_dir / target_rel)
        copied_files[source] = target_rel.as_posix()
        manifest_rows.append({"source": source, "relative_path": target_rel.as_posix(), "kind": "csv"})

        for pred_dir in candidate_prediction_paths(src_path):
            pred_key = str(pred_dir.resolve())
            if pred_key in copied_prediction_dirs:
                continue
            copied_prediction_dirs.add(pred_key)
            pred_rel = path_to_rel_copy(pred_dir.resolve())
            copy_any(pred_dir.resolve(), package_dir / pred_rel)
            manifest_rows.append(
                {
                    "source": str(pred_dir.resolve()),
                    "relative_path": pred_rel.as_posix(),
                    "kind": "predictions_dir",
                }
            )

    return copied_files, manifest_rows


def rewrite_workbook(
    package_dir: Path,
    sheet_names: list[str],
    frames: dict[str, pd.DataFrame],
    copied_files: dict[str, str],
) -> Path:
    out_workbook = package_dir / WORKBOOK_NAME
    rewritten = {}
    for sheet, frame in frames.items():
        out = frame.copy()
        for column in out.columns:
            if "source_csv" not in str(column):
                continue
            out[column] = out[column].fillna("").astype(str).map(
                lambda value: rewrite_source_value(value, copied_files)
            )
        rewritten[sheet] = out

    with pd.ExcelWriter(out_workbook, engine="openpyxl") as writer:
        for sheet in sheet_names:
            rewritten[sheet].to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            ws.freeze_panes = "A2"

    workbook = load_workbook(out_workbook)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                replacement = rewrite_source_value(cell.value, copied_files)
                if replacement != cell.value:
                    cell.value = replacement
    workbook.save(out_workbook)
    return out_workbook


def write_manifest(package_dir: Path, workbook: Path, copied_manifest_rows: list[dict[str, Any]]) -> None:
    manifest = {
        "workbook": WORKBOOK_NAME,
        "source_workbook": str(workbook),
        "package_dir": str(package_dir),
        "items": copied_manifest_rows,
    }
    (package_dir / "package_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main() -> None:
    args = parse_args()
    workbook = Path(args.workbook).expanduser().resolve()
    if not workbook.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    package_dir = Path(args.out_dir).expanduser().resolve()
    ensure_clean_dir(package_dir, args.overwrite)

    sheet_names, frames, source_values = gather_source_values(workbook)
    copied_files, copied_manifest_rows = package_sources(package_dir, source_values)
    out_workbook = rewrite_workbook(package_dir, sheet_names, frames, copied_files)
    write_manifest(package_dir, workbook, copied_manifest_rows)

    print(f"Packaged workbook to {out_workbook}")
    print(f"Copied {sum(1 for item in copied_manifest_rows if item['kind'] == 'csv')} CSV files")
    print(f"Copied {sum(1 for item in copied_manifest_rows if item['kind'] == 'predictions_dir')} prediction directories")


if __name__ == "__main__":
    main()
